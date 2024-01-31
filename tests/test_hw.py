import csv
from openpyxl import load_workbook
from pypdf import PdfReader
from helpers.directory import EXTRACTED_RESOURCES


def test_xlsx_file():
    reader = load_workbook(EXTRACTED_RESOURCES + 'example_xlsx.xlsx')
    sheet = reader.active
    c = sheet['A4']
    assert sheet.cell(1, 3).value == 'Название ЛПУ'
    assert c.value == 'Калужская область'


def test_csv_file():
    with open(EXTRACTED_RESOURCES + 'example_csv.csv') as file:
        rows = csv.reader(file)
        new_rows = [''.join(row).split(';') for row in rows]
        assert new_rows[0] == ['Tаблица 1',]
        assert new_rows[1][4] == 'Режим работы'
        assert new_rows[3][1] == 'Обнинск'


def test_pdf_file():
    with open(EXTRACTED_RESOURCES + 'example_pdf.pdf', 'rb') as file:
        reader = PdfReader(file)
        search_maloyaroslavets = any('Малоярославец' in page.extract_text() for page in reader.pages)
        assert search_maloyaroslavets, 'Малоярославец не найден в PDF файле'
